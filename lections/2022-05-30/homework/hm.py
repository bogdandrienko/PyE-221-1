import openpyxl
from openpyxl import Workbook


workbook = openpyxl.load_workbook('Лист Microsoft Excel.xlsx')
worksheet = workbook.active

print(workbook)
print(type(workbook))

value = worksheet.cell(2, 6).value
print(value)
print(type(value))

row1_arr = []
row2_arr = []
row_arr = []

# for i in range(1, worksheet.max_column+1):
#     row1_arr.append(worksheet.cell(1, i).value)

for i in range(1, 1000):
    value1 = worksheet.cell(1, i).value
    value2 = worksheet.cell(2, i).value
    # if value1 and value2:
    if value1 is not None and value2 is not None:
        row_arr.append([value1, value2])
    else:
        print('массивы закончились')

print(row_arr)
print(type(row_arr))

# for i in range(1, 1000):
#     value = worksheet.cell(2, i).value
#     if value:
#         row2_arr.append(value)
#
# print(row2_arr)
# print(type(row2_arr))

workbook_new = Workbook()
worksheet_new = workbook_new.active

# worksheet_new["A1"] = "111111111"

# for first, second in row_arr:
#     print(f"first: {first}")
#     print(f"second: {second}")

for pair in row_arr:
    index = row_arr.index(pair) + 1
    print(pair)
    worksheet_new[f"A{index}"] = pair[1]
    worksheet_new[f"B{index}"] = pair[0]

workbook_new.save('new.xlsx')
