import openpyxl

wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

data2 = []
data4 = []
data6 = []
for i in range(1, ws.max_column + 1):
    data2.append(ws.cell(row=2, column=i).value)
    data4.append(ws.cell(row=4, column=i).value)
    data6.append(ws.cell(row=6, column=i).value)

print(data2)
print(data4)
print(data6)

set2 = set(data2)
set4 = set(data4)
set6 = set(data6)

print(set2)
print(set4)
print(set6)

set_intersection = set2.intersection(set4, set6)
print(set_intersection)

set_difference = set2.difference(set4, set6)
print(set_difference)

set_union = set2.union(set4, set6)
print(set_union)
