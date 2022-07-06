import openpyxl
from openpyxl.utils import get_column_letter


class Calc:
    def __init__(self, value1: float, value2: float):
        self.value1 = value1
        self.value2 = value2

        self.sum = value1 + value2

    def sum(self):
        return self.sum

    def multiply(self):
        return self.value1 * self.value2

    @staticmethod
    def static_multiply(value1, value2):
        return value1 * value2


obj = Calc(12, 1.5)  # __init__
print(obj.multiply())  # multiply

print(Calc.static_multiply(1.5, 20))


class MyClass:
    def __init__(self, name="_1", index=1, value="_A1"):
        self.name = name
        self.index = index
        self.value = value

    def get_values(self):
        return [self.name, self.index, self.value]


records = []
for i in range(1, 1000 + 1):
    records.append(
        MyClass(
            name=f"_{i}",
            index=i,
            value=f"_A{i}",
        )
    )
print(records)

for i in records:
    print(i.get_values())

# Read all data from .xlsx
book = openpyxl.load_workbook("excel_data/tab.xlsx")
sheet = book.active
global_rows = []
for row in range(2, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)


class Record:
    def __init__(self, row: list):
        self.tovar = str(row[0])
        self.gruppa = str(row[1])
        self.postavshick = str(row[2])
        date_post = str(row[3]).split(" ")[0]  # '2010-12-4'
        year = date_post.split('-')[0]
        month = int(date_post.split('-')[1])  # '09' -> 9
        if int(month) < 10:
            month = f"0{month}"
        day = int(date_post.split('-')[2])
        if int(day) < 10:
            day = f"0{day}"
        self.date_post = str(f"{year}-{month}-{day}")
        self.region = str(row[4])
        try:
            self.prodashi = int(row[5])
        except Exception as error:
            print(error)
            self.prodashi = 0
        try:
            self.sbit = int(row[6])
        except Exception as error:
            print(error)
            self.sbit = 0
        pribil = str(row[7])
        if pribil.lower() == "да":
            self.pribil = True
        else:
            self.pribil = False


book_new = openpyxl.Workbook()
sheet_new = book_new.active
records = []
for row in global_rows:
    records.append(Record(row=row))

print(records)

row_index = 2
for row in records:
    sheet_new[f"A{row_index}"] = str(row.tovar)
    sheet_new[f"B{row_index}"] = str(row.prodashi)
    sheet_new[f"C{row_index}"] = str(row.postavshick)
    sheet_new[f"D{row_index}"] = str(row.pribil)
    row_index += 1
book_new.save('excel_data/new_new_table.xlsx')
