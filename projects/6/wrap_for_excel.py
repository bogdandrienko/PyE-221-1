import openpyxl


class MyExcel:
    def __init__(self, filename=None):
        if filename is None:
            self.workbook = MyExcel.create_new_workbook()
        else:
            self.workbook = MyExcel.load_workbook(filename=filename)
        self.worksheet = self.workbook.active
        self.name = filename

    def read_all_data(self, sheetname = '') -> list[list]:
        # TODO тут нужно дописать чтение определенного листа
        arr = []
        for row in range(1, self.worksheet.max_row + 1):
            arr_tmp = []
            for col in range(1, self.worksheet.max_column + 1):
                val = self.worksheet.cell(row=row, column=col).value
                if val is None:
                    val = ""
                arr_tmp.append(val)
            arr.append(arr_tmp)
        return arr

    # def __str__(self):
    #     return self.read_all_data()

    def fill_sheet_from_matrix(self, arr: list[list]) -> None:
        for i in range(1, len(arr) + 1):
            for j in range(1, len(arr[i-1]) + 1):
                self.worksheet.cell(row=i, column=j, value=arr[i-1][j-1])

    def save(self, name=None) -> None:
        if name is None:
            self.workbook.save(self.name)
        else:
            self.workbook.save(name)

    @staticmethod
    def load_workbook(filename: str) -> openpyxl.Workbook:
        return openpyxl.load_workbook(filename)

    @staticmethod
    def create_new_workbook() -> openpyxl.Workbook:
        return openpyxl.Workbook()


wb1 = MyExcel(filename='data.xlsx')
print(wb1)

arr_tmp = [[0, 4, 5], [3, 2, 6]]
wb1.fill_sheet_from_matrix(arr_tmp)
print(wb1)
wb1.save(name='data_new.xlsx')
