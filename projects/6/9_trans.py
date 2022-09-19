import openpyxl
import json

# workbook = openpyxl.load_workbook("data1.xlsx")
# worksheet = workbook.active
# sell1 = worksheet["B2"].value
# print(sell1)
# list1 = []
# for i in "ABCD":
#     list1.append(worksheet[f"{i}2"].value)
# print(list1)
# new_workbook = openpyxl.Workbook()
# new_worksheet = new_workbook.active
# index = 1
# for j in list1:
#     new_worksheet[f"B{index}"] = j
#     index+=1
#
# new_worksheet["C3"] = 666
# new_workbook.save("data2.xlsx")


# workbook = openpyxl.load_workbook("data1.xlsx")
# worksheet = workbook.active
# sell1 = worksheet["B2"].value
# print(sell1)
#
# list_external =[]
# for j in range(1, 5):
#     list_internal = []
#     for i in "ABCD":
#         list_internal.append(worksheet[f"{i}{j}"].value)
#     list_external.append(list_internal)
# print(list_external)
#
# key1 = list_external[0][0]
# key2 = list_external[0][1]
# key3 = list_external[0][2]
# key4 = list_external[0][3]
# print(key1, key2, key3, key4)
#
#
#
#
# dict1 = {key1: list_external[1][0], key2: list_external[1][1], key3: list_external[1][2], key4: list_external[1][3]}
# print(dict1)
# with open("data1.json", mode="w") as file:
#     json.dump(dict1, file)

id = -1
name = 3
surname = 3
age = 4
salary = 5
