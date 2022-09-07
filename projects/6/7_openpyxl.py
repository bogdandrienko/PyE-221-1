from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
import openpyxl
import datetime
import random

# TODO CREATE AND WRITE ##################################################################

workbook = Workbook()  # создание инстанса - экземпляра класса
worksheet = workbook.active  # активация 1 рабочего листа
arr1 = [1, 151513, 12, 66, "rgwrg", datetime.datetime.now()]
arr2 = [1, 5, 12, 1124124, "rgwrg", datetime.datetime.now()]
arr3 = [1, 1241, 12, 66, "rgwrg", datetime.datetime.now()]
worksheet['B2'] = 666  # присвоение ячейки с координатами значения

# for item in set(arr1):
#     worksheet[f'A{arr1.index(item)+1}'] = item

# for elem in "ABCD":
#     index = "ABCD".index(elem)
#     worksheet[f'{elem}1'] = arr1[index]
#     worksheet[f'{elem}2'] = arr2[index]
#     worksheet[f'{elem}3'] = arr3[index]

words = ["Мяч", "Столб", "Тетрадь", "openpyxl", "worksheet", "Worksheet"]

massiv1 = [x for x in range(1, 5 + 1)]
massiv2 = [random.choice(words) for x in range(1, 5 + 1)]
massiv_s_massivami = [massiv1, massiv2, massiv1, [random.choice(words) for x in range(1, 5 + 1)]]
# print(list1)
# print(list2)
# print(random.choice(words))
# print(massiv_s_massivami)

# for row_i in range(0, len(massiv_s_massivami)):
#     massiv = massiv_s_massivami[row_i]
#
#     for col_i in range(0, len(massiv)):
#         znachenie = massiv[col_i]
#
#         print(f"row: {row_i+1}, col: {col_i+1} value: {znachenie}")
#         worksheet.cell(row=row_i+1, column=col_i+1, value=znachenie)

row_i = 0
for massiv in massiv_s_massivami:
    row_i += 1

    col_i = 0
    for znachenie in massiv:
        col_i += 1

        # print(f"row: {row_i}, col: {col_i} value: {znachenie}")
        worksheet.cell(row=row_i, column=col_i, value=znachenie)


def get_password(length: int, elem="") -> str:
    if length <= 0 or len(elem) <= 0:
        raise Exception
    password = ""
    while len(password) < length:
        password += random.choice(elem)
    return password


# print(get_password(length=8, elem="ABNFGKFGOIWNPV1234567890!=-"))

# workbook.save("example.xlsx")  # сохранение рабочей книги

# TODO READ ##################################################################

# workbook = openpyxl.load_workbook(filename="data.xlsx")
# worksheet = workbook.active

# val1 = worksheet['E1'].value
# print(val1)
# print(type(val1))

# external_matrix = []
# for i in range(1, 4+1):
#     internal_matrix = []
#     for j in range(1, 4+1):
#         # internal_matrix.append(worksheet[f'{get_column_letter(j)}{i}'].value)
#         internal_matrix.append(worksheet.cell(row=i, column=j).value)
#     external_matrix.append(internal_matrix)

# print(external_matrix)

# row1 = [worksheet.cell(row=x, column=1).value for x in range(1, 10)]
# print(row1)

# worksheet['A1'] = 111
# worksheet.cell(row=1, column=1, value=111)

# TODO READ AND UPDATE ##################################################################

workbook = Workbook()
# workbook.create_sheet()
worksheet = workbook.active
font = Font(name='Tahoma', size=16, bold=True,
            italic=False, vertAlign=None, underline='none',
            strike=False, color='FF0000FF')
worksheet['B2'].font = font
worksheet['B2'] = 'Python'

workbook.save("data.xlsx")

# workbook = openpyxl.load_workbook('document.xlsx')
# workbook.template = True
# workbook.save('document_template.xltx')
