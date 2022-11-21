import datetime
import random

from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.hashers import make_password
from django.contrib.auth.models import User, Group, update_last_login
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django_twitter_app import models
import openpyxl
from openpyxl.utils import get_column_letter


# Create your views here.


class HomeView(View):  # TODO контроллер класс
    template_name = 'django_twitter_app/home.html'

    def get(self, request: HttpRequest) -> HttpResponse:
        context = {}
        # return HttpResponse(content=b"<h1>Hello World</h1>")
        # return JsonResponse(data={"response": 'res'}, safe=True)
        return render(request, 'django_twitter_app/home.html', context=context)

    def post(self, request: HttpRequest) -> HttpResponse:
        context = {}
        # return HttpResponse(content=b"<h1>Hello World</h1>")
        # return JsonResponse(data={"response": 'res'}, safe=True)
        return render(request, 'django_twitter_app/home.html', context=context)


def home_view(request: HttpRequest) -> HttpResponse:  # TODO контроллер функция
    context = {}
    # return HttpResponse(content=b"<h1>Hello World</h1>")
    # return JsonResponse(data={"response": 'res'}, safe=True)
    return render(request, 'django_twitter_app/home.html', context=context)


def register(request: HttpRequest) -> HttpResponse:
    #

    if request.method == "GET":
        context = {}
        return render(request, 'django_twitter_app/register.html', context=context)
    elif request.method == "POST":

        # TODO получить с формы данные
        first_name = request.POST.get('first_name', "")
        last_name = request.POST.get('last_name', "")
        username = request.POST.get('username', None)
        password1 = request.POST.get('password1', "")
        password2 = request.POST.get('password2', "")

        if password1 and password1 != password2:
            raise Exception("пароли не совпадают!")
        if username and password1:
            User.objects.create(
                first_name=first_name,
                last_name=last_name,
                username=username,
                password=make_password(password1),
            )
            return redirect(reverse('django_twitter_app:login', args=()))  # name=
        else:
            raise Exception("данные не заполнены!")


def login_f(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        context = {}
        return render(request, 'django_twitter_app/login.html', context=context)
    if request.method == "POST":
        username = request.POST.get('username', "").strip()
        password = request.POST.get('password', "").strip()
        # print(f"username: {username}", password)
        # str1 = ""  # 0 False None "" '' [] (,)
        # if str1:
        #     print("Правда")
        # else:
        #     print("Ложь")
        # if username is not False and password is not False:
        #     pass
        # if len(username) > 0 and len(password) > 0:
        #     pass

        if username and password:  # если оба выражения "правда"
            # User.objects.get(username=username) = authenticate(username=username, password=password)  # success
            # None = authenticate(username=username, password=password)  # fail
            user_obj = authenticate(username=username, password=password)  # проверка на валидность
            if user_obj:
                if user_obj.is_active is False:
                    raise Exception("Ваш аккаунт забанен!")
                context = {}
                login(request, user_obj)
                # login(user_obj)  # вход
                # logout(request)
                # update_last_login(sender=None, user=user_obj)
                return redirect(reverse('django_twitter_app:home', args=()))
            else:
                raise Exception("данные не совпадают!")
        else:
            raise Exception("данных нет!")


def logout_f(request: HttpRequest) -> HttpResponse:
    logout(request)
    # context = {}
    # return render(request, 'django_twitter_app/home.html', context=context)
    return redirect(reverse('django_twitter_app:login', args=()))


def post_create(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        context = {}
        return render(request, 'django_twitter_app/post_create.html', context=context)
    elif request.method == "POST":
        print("request: ", request)
        # print("request.data: ", request.data)
        print("request.POST: ", request.POST)
        print("request.GET: ", request.GET)
        print("request.META: ", request.META)

        title = request.POST.get('title', None)
        description = request.POST.get('description', "")
        post = models.Post.objects.create(
            user=request.user,
            title=title,
            description=description,
        )
        return redirect(reverse('django_twitter_app:post_list', args=()))
        # context = {"id": post.id}
        # return redirect(reverse('django_twitter_app:post_list', args=(context, )))


def post_update(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "GET":
        post = models.Post.objects.get(id=pk)  # primary key
        context = {"post": post}
        return render(request, 'django_twitter_app/post_update.html', context=context)
    elif request.method == "POST":
        print("request: ", request)
        # print("request.data: ", request.data)
        print("request.POST: ", request.POST)
        print("request.GET: ", request.GET)
        print("request.META: ", request.META)

        # TODO получить с формы данные
        title = request.POST.get('title', None)
        description = request.POST.get('description', "")

        # TODO получить нужный объект с базы данных
        post = models.Post.objects.get(id=pk)

        # TODO обновить объект в базе данных
        post.title = title
        post.description = description
        post.save()

        #####################
        # title = request.data.get('title', [])  # {"title": [{}, {}, {}, {}], "name": "Python"}

        ####################

        # TODO перенаправить
        return redirect(reverse('django_twitter_app:post_detail', args=(pk,)))
        # context = {"id": post.id}
        # return redirect(reverse('django_twitter_app:post_list', args=(context, )))


class CustomPaginator:
    @staticmethod
    def paginate(object_list: any, per_page=5, page_number=1):
        # https://docs.djangoproject.com/en/4.1/topics/pagination/
        paginator_instance = Paginator(object_list=object_list, per_page=per_page)
        try:
            page = paginator_instance.page(number=page_number)
        except PageNotAnInteger:
            page = paginator_instance.page(number=1)
        except EmptyPage:
            page = paginator_instance.page(number=paginator_instance.num_pages)
        return page


def post_list(request: HttpRequest) -> HttpResponse:
    # title  descr
    # 111111 222222
    # 22222 33333
    posts = models.Post.objects.all()  # filter order_by DjangoORM | SQLAlchemy(Flask) | Pydantic (FastApi)

    # todo пагинация ч1
    selected_page_number = request.GET.get('page', 1)
    selected_limit_objects_per_page = request.GET.get('limit', 2)
    # todo пагинация ч1

    if request.method == "POST":
        selected_page_number = 1
        selected_limit_objects_per_page = 9999
        # todo поиск
        search_by_title = request.POST.get('search', None)
        if search_by_title is not None:
            posts = posts.filter(title__contains=str(search_by_title))  # title__icontains
        # todo поиск

        # todo фильтрация
        filter_by_user = request.POST.get('filter', None)
        if filter_by_user is not None:
            posts = posts.filter(user=User.objects.get(username=filter_by_user))
        # todo фильтрация

    # todo пагинация ч2
    page = CustomPaginator.paginate(
        object_list=posts, per_page=selected_limit_objects_per_page, page_number=selected_page_number
    )
    # todo пагинация ч2

    context = {"page": page, "users": User.objects.all()}
    return render(request, 'django_twitter_app/post_list.html', context=context)


def post_detail(request: HttpRequest, pk: int) -> HttpResponse:
    post = models.Post.objects.get(id=pk)
    comments = models.PostComment.objects.filter(article=post)

    selected_page_number = request.GET.get('page', 1)
    selected_limit_objects_per_page = request.GET.get('limit', 2)
    page = CustomPaginator.paginate(
        object_list=comments, per_page=selected_limit_objects_per_page, page_number=selected_page_number
    )

    # TODO ratings
    ratings = models.PostRating.objects.all().filter(article=post)  # queryset
    likes = ratings.filter(status=True)
    dislikes = ratings.filter(status=False)
    context = {"post": post, "page": page, "ratings": [likes.count(), dislikes.count(), ratings.count()]}
    return render(request, 'django_twitter_app/post_detail.html', context=context)


def post_delete(request: HttpRequest, pk: int) -> HttpResponse:
    # models.Post.objects.get(id=pk).delete()
    post = models.Post.objects.get(id=pk)
    post.delete()
    return redirect(reverse('django_twitter_app:post_list', args=()))


def post_rating(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "POST":
        # TODO ratings

        post = models.Post.objects.get(id=pk)

        status = request.POST.get("status", None)
        if status is None:
            raise Exception("No status")
        elif status == "like":
            try:
                rating = models.PostRating.objects.get(article=post, user=request.user)
            except Exception as error:
                rating = models.PostRating.objects.create(
                    user=request.user,
                    article=post
                )
            if rating.status is False:
                rating.status = True
                rating.save()
            else:
                pass
        elif status == "dislike":
            try:
                rating = models.PostRating.objects.get(article=post, user=request.user)
            except Exception as error:
                rating = models.PostRating.objects.create(
                    user=request.user,
                    article=post
                )
            if rating.status is True:
                rating.status = False
                rating.save()
            else:
                pass
        else:
            raise Exception("Error status")
        return redirect(reverse('django_twitter_app:post_detail', args=(pk,)))


def post_pk_view(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "GET":
        # post_list = models.Post.objects.all()
        # print(f"post_list: {post_list}")
        # context = {"post_list": post_list}
        context = {}
        return render(request, 'django_twitter_app/post_detail.html', context=context)
    context = {}
    # return HttpResponse(content=b"<h1>Hello World</h1>")
    # return JsonResponse(data={"response": 'res'}, safe=True)
    return render(request, 'django_twitter_app/post_list.html', context=context)


def post_comment_create(request: HttpRequest, pk: int) -> HttpResponse:
    if request.method == "POST":
        text = request.POST.get('text', None)
        post = models.Post.objects.get(id=pk)  # определить, к какой статье создали комментарий
        models.PostComment.objects.create(
            user=request.user,
            article=post,
            text=text,
            # date_time=timezone.now(), # у нас стоит default
        )
        return redirect(reverse('django_twitter_app:post_detail', args=(pk,)))


def post_comment_delete(request: HttpRequest, pk: int) -> HttpResponse:
    comment = models.PostComment.objects.get(id=pk)
    pk = comment.article.id
    comment.delete()
    return redirect(reverse('django_twitter_app:post_detail', args=(pk,)))


def test(request: HttpRequest) -> HttpResponse:
    status = request.GET.get("status", None)
    return HttpResponse(f"test: <h1>{status}</h1>")


def test_by_filter(request: HttpRequest, filter: str) -> HttpResponse:
    return HttpResponse(f"test_by_filter: <h1>{filter}</h1>")


def notification_create(request: HttpRequest) -> HttpResponse:
    if request.method == "POST":
        title = request.POST.get('title', None)
        description = request.POST.get('description', "")
        # post = models.Post.objects.create(
        #     user=request.user,
        #     title=title,
        #     description=description,
        # )

        with open("static/media/admin/notification.txt", "a", encoding="utf-8") as opened_file:
            opened_file.write(f"\n{title} | {timezone.now()}")

        return redirect(reverse('django_twitter_app:post_list', args=()))
        # context = {"id": post.id}
        # return redirect(reverse('dja


def export_users(request: HttpRequest) -> HttpResponse:
    # users = User.objects.all().filter()
    users = User.objects.filter(is_active=True)  # READ

    path = "static/media/temp/"
    filename = f"new{random.randint(1, 10000)}{datetime.datetime.now().strftime('%m-%d-%Y, %H-%M-%S')}.xlsx"
    # new_excel_workbook = openpyxl.Workbook()
    already_exist_excel_workbook = openpyxl.load_workbook(filename="static/media/admin/export_users.xlsx")
    new_excel_workbook = already_exist_excel_workbook
    worksheet = new_excel_workbook.active

    # values = [1, 2, 3]
    # for index in range(0, len(values)):
    #     value = values[index]  # лишний код и расчёты !

    # getter = worksheet["A1"].value  # getter
    # worksheet["A1"] = "123123"  # setter
    # row_i = 0  # лишний код и расчёты !
    for index, value in enumerate(["username", "password", "email", "is_staff", "is_superuser"], 1):
        worksheet.cell(1, index, value)  # setter

    for row_index, user in enumerate(users, 2):

        #  0  1  2  3
        # [1, 2, 2, 3]
        # 0  1   1

        # row_i_ = users.index(user)  # берёт первое совпадение!!! если значения не уникальные нельзя!!!
        # row_i += 1  # лишний код и расчёты !
        username = user.username
        password = user.password
        email = user.email
        is_staff = user.is_staff
        is_superuser = user.is_superuser
        cols = [username, password, email, is_staff, is_superuser]
        for col_index, value in enumerate(cols, 1):
            # getter = worksheet.cell(1, 1).value  # getter
            worksheet.cell(row_index, col_index, value)  # setter

    new_excel_workbook.save(path + filename)
    return HttpResponse("<h1>Success</h1>")


# HTTP/IP - протокол транспортного уровня
# UDP - не создаёт соединение (просто кидает данные в одну сторону)


def create_users(request: HttpRequest) -> HttpResponse:  # MVTemplate - html
    if request.method == "GET":
        context = {}
        return render(request, 'django_twitter_app/create_users.html', context=context)
    if request.method == "POST":
        # uploaded_excel_file = request.GET.get('excel', None)  # https://hh.ru/search/vacancy? @text=python@ &from=suggest_post&area=154
        # uploaded_excel_file = request.POST.get('excel', None)  # {"name": "Bogdan"...}
        uploaded_excel_file = request.FILES.get('excel', None)  # <image> / <file>

        # todo read excel
        already_exist_excel_workbook = openpyxl.load_workbook(uploaded_excel_file)
        # already_exist_excel_workbook.active = already_exist_excel_workbook['Лист 2']
        worksheet = already_exist_excel_workbook.active

        class Profile:
            def __int__(self, username_: str, password_: str):
                self.username_ = username_
                self.password_ = password_
                pass

        rows = []  # {"username": "Bogdan"...}
        for i in range(1, worksheet.max_row):
            row = []
            for j in range(1, worksheet.max_column):
                cell = worksheet[f"{get_column_letter(j)}{i}"].value
                row.append(cell)
            rows.append(row)
        print(rows)

        def prettify(value_str: str, length: int) -> str:
            while len(value_str) < length:
                value_str += " "
            return value_str

        # for row in rows:
        #     matrix = []
        #     for index, value in enumerate(row):
        #         value = prettify(value_str=str(value), length=30)
        #         matrix.append(value)
        #     print(matrix)
        # todo read excel

        # todo create users
        for row in rows:
            try:
                username = row[0]
                if len(username) < 2:
                    continue
                password = row[1]
                email = row[2]
                surname = str(row[3]).strip()
                is_staff = True if str(row[4]).strip().lower() == "true".lower() else False
                user = User.objects.create(
                    username=username,
                    password=make_password(password),
                    email=email,
                    first_name=surname,
                    is_staff=is_staff,
                    is_superuser=False,
                )
            except Exception as error:
                print(error)
            # user = User.objects.create_user()
            # user = User.objects.create_superuser()

        return redirect(reverse('django_twitter_app:create_users', args=()))


# REST api - mobile (kotlin - java / swift - object c) web(react / angular / vue, next, nuxt...) desktop (electron,
# pyqt6, qt...) console ()

from django.views.decorators.csrf import csrf_exempt


@csrf_exempt
def test2(request: HttpRequest) -> HttpResponse:
    if request.method == "GET":
        print(request.GET)
        print(request.POST)
        print(request.FILES)
        return HttpResponse("GET method")
    if request.method == "POST":
        try:
            # print(request.META)
            print(request.GET)
            print(request.POST)
            print(request.FILES)
        except Exception as error:
            return JsonResponse(data={"message": str(error)}, safe=False, status=400)
            # return HttpResponse(str(error))
        else:
            return JsonResponse(data={"message": "successful"}, safe=False, status=200)
            # return HttpResponse("successful")
