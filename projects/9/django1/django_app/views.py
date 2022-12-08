import os
import sqlite3

import openpyxl
from django.http import HttpRequest, HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse

from django_app import models


def index(request: HttpRequest) -> HttpResponse:

    data_arr = tuple(({"id": x, "name": f"{x}", "amount": x**2} for x in range(1, 3)))
    print(data_arr, type(data_arr))

    context = {"data_arr": data_arr}

    return render(request, "index.html", context)


def create(request: HttpRequest) -> HttpResponse:

    label = str(request.POST.get("name", ""))
    amount = int(request.POST.get("amount", 0))
    not_bubble_price = int(request.POST.get("first_cost", 0))
    bubble_percentage = int(request.POST.get("percent", 0))

    final_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
    vat_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
    overall = final_price * amount

    models.Product.objects.create(
        label=label,
        amount=amount,
        not_bubble_price=not_bubble_price,
        bubble_percentage=bubble_percentage,
        final_price=final_price,
        vat_price=vat_price,
        overall=overall,
    )

    # TODO создание в базе данных:
    # SQL

    return redirect(reverse("index", args=()))


def delete(request: HttpRequest, pk: int) -> HttpResponse:

    print(pk, type(pk))

    # TODO удаление из базы данных:
    # ORM
    # SQL

    # ProductModel.objects.get(id=pk).delete()

    return redirect(reverse("index", args=()))


def import_data(request: HttpRequest) -> HttpResponse:
    """Import data from xlsx to sqlite"""

    excel_file = request.FILES["excel_file"]
    if excel_file:
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook.active  # sheet = file_to_read["Лист1", "Sheet2", "Sheet3"]

        matrix = []
        for row in range(2, worksheet.max_row + 1):
            data = []
            for column in range(1, worksheet.max_column + 1):
                value = worksheet.cell(row, column).value

                if value is None or value == "-" or value == "":
                    value = 0

                if column == 1:
                    value = str(value)
                else:
                    value = int(str(value).strip())

                data.append(value)
            matrix.append(data)

        print(matrix)
        for product in matrix:

            label = str(product[0])
            amount = int(product[1])
            not_bubble_price = int(product[2])
            bubble_percentage = int(product[3])

            final_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
            vat_price = not_bubble_price + (round(not_bubble_price * 12 / 100, 0))
            overall = final_price * amount

            models.Product.objects.create(
                label=label,
                amount=amount,
                not_bubble_price=not_bubble_price,
                bubble_percentage=bubble_percentage,
                final_price=final_price,
                vat_price=vat_price,
                overall=overall,
            )

    # TODO запись в базу данных:
    # SQL

    return redirect(reverse("index", args=()))

    # prj_dir = os.path.dirname(os.path.dirname(os.path.abspath(file)))  # Получение ДБ, не уверен

    # # 1.Сonnecting to the database

    # # file

    # connect = sqlite3.connect(prj_dir + "/")

    # cursor = connect.cursor()

    # # 2. Working with xlsx file

    # for row in range(2, sheet.max_row + 1):

    #     data = []

    #     for col in range(1, 5):
    #         value = sheet.cell(row, col).value

    #         data.append(value)

    #     # 3. Writing to the database and closing the connection

    #     cursor.execute("INSERT INTO ... VALUES (?, ?, ?, ?);", (data[0], data[1], data[2], data[3]))

    # # Save change
    # connect.commit()
    # # Close connection
    # connect.close()


def export_data(request: HttpRequest) -> HttpResponse:
    products = models.Product.objects.all()
    print(products, type(products))

    for row_index, product in enumerate(products, 1):
        label = product.label
        amount = product.amount
        not_bubble_price = product.not_bubble_price
        bubble_percentage = product.bubble_percentage
        final_price = product.final_price
        vat_price = product.vat_price
        overall = product.overall

        # for column_index, value in enumerate(row, 1):
        #     print(f"{row_index} {column_index} {value}")

        print(f"{label}: [{amount}]  = {overall}")

    return redirect(reverse("index", args=()))
