import datetime
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import Workbook


dir_name = 'dist'
files = []

# читаем все файлы из заданной директории(папки), только xlsx
for filename in os.listdir(dir_name):  # os.listdir - принимает в себя директорию откуда нужно вывести все имена файлов
    # print(filename.upper())  # делает все буквы заглавными
    # print(filename.lower())  # делает все буквы прописными
    # print(filename.capitalize())  # делает все буквы как в предложении(первая заглавная, остальные маленькие)
    # print(filename.strip())  # очищает пробелы, табуляции и знаки переноса со строки по краям

    filename1 = filename.split(sep='.')  # метод для строки, который режет её по сепаратору(набор символов)
    # print(filename)
    # print(type(filename))
    # print(filename1)
    # print(type(filename1))

    if filename1[-1] == 'xlsx' or filename1[-1] == 'xls':  # учли, что формат файла может быть устаревшим
        files.append(filename)
    else:
        pass
print(files)

data = []

for file in files:  # проходим циклом по массиву с нужными именами файлов

    # конкатенация строк или сложение строк
    file_name1 = dir_name + '/' + file
    # print(file_name1)

    try:
        workbook = openpyxl.load_workbook(file_name1)  # принимает имя файла и открывает его как рабочую книгу
    except:
        # происходит выполнение код, когда код в блоке 'try' вызвал исключение(ошибку)
        print(f'Закройте файл "{file}"!')
        continue

    worksheet = workbook.active
    max_row = worksheet.max_row
    # print(max_row)
    max_column = worksheet.max_column + 1
    # print(max_column)

    dict1 = {}  # создаём пустой словарь(тип данных ключ-значение)

    city = file.split('.')[-2].split(' ')[-1]
    # print(f"Имя файла: {city}")
    if city.isdigit():
        # пропустить итерацию цикла
        continue

    dict1["Город"] = city

    # dict1["Общий охват"] = worksheet.cell(row=6, column=3).value



    #############################################################

    # тут будет логика обработки excel

    start = str(worksheet["C6"].value).split(":")[0].split("(")[1]
    # print(f"start: {start}")
    end = str(worksheet["C6"].value).split(":")[1].split(sep=")")[0]
    # print(f"end: {end}")

    # print(tuple(worksheet[start:end]))

    oxvat = ''
    for i in tuple(worksheet[start:end]):
        for j in i:
            if j.value:
                oxvat = j.value
            else:
                print(f'ВНИМАНИЕ город {city} данные не заполнены!')
                oxvat = 0

    total = worksheet.cell(row=5, column=2).value
    if total:
        pass
    else:
        total = 0
    dict1["Всего мероприятий"] = total

    dict1["Общий охват"] = oxvat

    vstr = worksheet["D5"].value
    if not vstr:
        vstr = 0
    dict1["Всего встреч"] = vstr

    #############################################################

    # worksheet[f'{col}{row}'] = str(name)

    data.append(dict1)
    # print("\n\n\n")
print(data)

# объединяем все словари в один сводный excel-файл

new_dir_name = 'result'
# ИСКЛЮЧЕНИЯ
# Exception - исключение т.е. критическая ошибка
try:
    # происходит ПОПЫТКА выполнить какой-то код, который может вызвать ошибку

    # путь относительно текущего скрипта
    os.mkdir(dir_name + "/" + new_dir_name)
    # connection - объект в памяти, который непрерывно обменивается данными
except:
    # происходит выполнение код, когда код в блоке 'try' вызвал исключение(ошибку)
    print('Папка уже существует!')
finally:
    # происходит выполнение кода, безотносительно удачного или неудачного выполнения
    # тут нужно закрыть соединение с базой данных или файл для чтения/записи
    pass


workbook2 = Workbook()
worksheet2 = workbook2.active

titles = ["Город", 'мероприятий', 'Охват', 'Встреча']
index = 1
for title in titles:
    worksheet2[f"{get_column_letter(index)}1"] = title
    index += 1

total_vstr = 0
total_obxv = 0
total_mer = 0

extra_index = 2
for row in data:
    print(row)
    print(type(row))

    # проходим по словарю циклом for
    # print(row.values())  # возвращает из словаря значения
    # print(row.keys())  # возвращает из словаря ключи
    # print(row.items())  # возвращает из словаря ключи и значения

    # for col in row.items():
    #     print(f'key = {col[0]}')
    #     print(f'value = {col[1]}')

    index = 1
    for key, value in row.items():
        # print(f'key = {key}')
        # print(f'value = {value}')

        if key == 'Всего встреч':
            total_vstr += int(value)

        if key == 'Общий охват':
            total_obxv += int(value)

        if key == 'Всего мероприятий':
            total_mer += int(value)

        worksheet2[f'{get_column_letter(index)}{extra_index}'] = value
        index += 1
    extra_index += 1

titles = ["Итого", total_mer, total_obxv, total_vstr]
index = 1
for title in titles:
    worksheet2[f"{get_column_letter(index)}{extra_index}"] = title
    index += 1


workbook2.save(f"{dir_name}/{new_dir_name}/result.xlsx")
