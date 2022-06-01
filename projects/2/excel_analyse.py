from tkinter import *  # может произойти коллизия имён (перезапись предыдущего имени последующим)
from tkinter import ttk
import tkinter
from tkinter import messagebox

import openpyxl


def cancel():
    pass


# определение(создание) функции
def get_result():
    col_from_user = col_label.get()
    print(f'колонки: {col_from_user}')
    elem_from_user = elem_label.get()
    print(f'элемент: {elem_from_user}')
    ############################################################

    file_name = 'temp/sample_example_new.xlsx'
    workbook = openpyxl.load_workbook(file_name)

    # берём активную страницу из рабочей книги
    worksheet = workbook.active

    # берём последнюю строку в excel-файле
    max_row = worksheet.max_row
    print(max_row)
    # берём последнюю колонку в excel-файле
    max_column = worksheet.max_column
    print(max_column)

    # записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
    external_array = []

    # наполнение внешнего массива
    for row in range(1, max_row + 1):
        internal_array = []

        # наполнение внутреннего массива
        for col in range(1, max_column + 1):
            value = worksheet.cell(row=row, column=col).value
            if value is None:
                value = ''
                # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
                # break  # (прекратить) оператор позволяет полностью остановить цикл
            internal_array.append(value)
        # завершение наполнение внутреннего массива

        if len(internal_array) < 1:
            continue

        external_array.append(internal_array)
    # завершение наполнение внешнего массива

    print(external_array)

    # '1, 2, 6' -> [1, 2, 6]
    col_list = []
    print(col_from_user)
    print(type(col_from_user))
    for i in col_from_user.split(","):  # метод, который вызывается на строках (режет строку на массив)
        # "1, 2, 3 , 5".split(",") -> ["1", " 2", " 3 ", " 5"]
        i = str(i).strip()  # " 1" / "2 " -> "1" / "2"
        i = int(i)  # "1" -> 1
        col_list.append(i)
        print(i)
        print(type(i))
    print(col_list)

    # elem_from_user
    count = 0
    for i_new in external_array:
        for j_new in i_new:
            if elem_from_user == j_new:
                count = count + 1

    ###########################################################

    # тут мы уже должны посчитать количество вхождений
    result = count
    result_label.config(text=f"{result}")


# инициализация инстанса - создание объекта ткинтер
root = Tk()

# создание главного окна
frm = ttk.Frame(root, padding=100)
root.title("Анализ данных")
root.geometry("640x480")
frm.grid()

# результат
result_label = ttk.Label(frm, text="Количество вхождений: ")
result_label.grid(column=4, row=0)

# колонки
col_label_var = tkinter.StringVar()  # переменная, которая хранит колонки
col_label = ttk.Entry(textvariable=col_label_var)
col_label.grid(column=1, row=3)

# элемент
elem_label_var = tkinter.StringVar()  # переменная, которая хранит элемент
elem_label = ttk.Entry(textvariable=elem_label_var)
elem_label.grid(column=5, row=3)

# кнопка стоп
Button(text="stop",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=cancel,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=0, row=5)

# кнопка старт
Button(text="start",  # текст кнопки
       background="#555",  # фоновый цвет кнопки
       foreground="#ccc",  # цвет текста
       padx="20",  # отступ от границ до содержимого по горизонтали
       pady="8",  # отступ от границ до содержимого по вертикали
       font="16",  # высота шрифта
       command=get_result,  # ОБЯЗАТЕЛЬНО ПЕРЕДАВАТЬ ССЫЛКУ НА ФУНКЦИЮ
       ).grid(column=2, row=5)

root.mainloop()

str1 = "124124|14121|1411234".split(sep="|")  # -> ["124124", "14121", "1411234"]
