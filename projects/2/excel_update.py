import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# чтение
# обрабатываем данные
# записываем в исходный файл (можно исходный файл удалять)


workbook = openpyxl.load_workbook("""temp/sample_example.xlsx""")

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row
print(max_row)
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column
print(max_column)

# записать всё в массив с массивами [["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"], ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]...]
external_array = []

# наполнение внешнего массива
for row in range(1, max_row + 1):
    internal_array = []

    # наполнение внутреннего массива
    for col in range(1, max_column + 1):
        value = worksheet.cell(row=row, column=col).value
        if value is None:
            value = ''
            # continue  # (продолжить) оператор позволяет пропустить эту итерацию цикла
            # break  # (прекратить) оператор позволяет полностью остановить цикл
        internal_array.append(value)
    # завершение наполнение внутреннего массива

    if len(internal_array) < 1:
        continue

    external_array.append(internal_array)
# завершение наполнение внешнего массива

print(external_array)

# index = 0
# while True:
#     index += 1
#     if index > 50:
#         break # (прекратить) оператор позволяет полностью остановить цикл

workbook_new = Workbook()
worksheet_new = workbook_new.active

# col_count = external_array[0]
print(
    f"col_count: {external_array}")  # [['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''], ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', ''],]
print(f"col_count: {external_array[0]}")  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
print(f"col_count: {external_array[0][1]}")  # 'Almaty'
print(f"col_count: {external_array[0][1][2:-2:1]}")  # 'ma'


def function_len_array(array):  # ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']
    return len(array)  # 5


# -2 -1 0 1 2

for row in range(0, function_len_array(external_array)):  # [0, 1, 2, 3, 4, ..., 1007]
    # print(f"col_count: {len(external_array[row])}")
    for col in range(0, function_len_array(external_array[row])):
        # [0, 1, 2, 3, 4, 5]  # external_array[row] == ['Almaty', 'Nur-Sultan', 'Taraz', 'Ekibastuz', '']

        worksheet[f'{get_column_letter(col + 1)}{row + 1}'] = external_array[row][col]
        if row == 0:
            worksheet[f'{get_column_letter(col + 1)}{row + 1}'].font = Font(bold=True)
        pass
    # print(external_array[row])
    pass

# for char in 'ABCDE':
# worksheet_new['A1'].font = Font(bold=True)
# worksheet_new['B1'].font = Font(bold=True)
# worksheet_new['C1'].font = Font(bold=True)
# worksheet_new['D1'].font = Font(bold=True)
# worksheet_new['E1'].font = Font(bold=True)


workbook.save('temp/sample_example_new.xlsx')

wb = Workbook()
ws = wb.active
ws['B2'] = "Hello1"
ws['B3'] = "Hello"
ws['B2'].font = Font(bold=True)
ws['B3'].font = Font(bold=True)
wb.save("temp/BoldDemo.xlsx")

# читаем первый файл, второй, третий... файлы
# складываем значение в каждом уникальном массиве
# создаём новый файл, который содержит в себе финальное значение (+ форматирование)
# как получить все имена файлов в папке

# форматирование: жирный, курсив, размер, capitalize, фон, границы, подчёркивание, формулы, объединение, подбор ширины
