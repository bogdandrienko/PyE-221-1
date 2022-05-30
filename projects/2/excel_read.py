# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_interval

file_name = "temp/sample_example.xlsx"

# загружаем в память уже существующий файл на диске
workbook = openpyxl.load_workbook(file_name)

# берём активную страницу из рабочей книги
worksheet = workbook.active

# берём последнюю строку в excel-файле
max_row = worksheet.max_row + 1
print(max_row)
print(type(max_row))
# берём последнюю колонку в excel-файле
max_column = worksheet.max_column + 1
print(max_column)
print(type(max_column))

index = 0
for i in range(1, max_row):

    # # получение значения с выбранной ячейки, где row - строка, column - колонка
    # # value = worksheet.cell(row=i, column=1).value
    #
    # # получение значения с выбранной ячейки, где в квадратных скобках координаты ячейки
    # value = worksheet[f"A{i}"].value
    #
    # # if value is not None:
    # #     pass
    # # if len(str(value)) >= 2:
    # if value:
    #     print(value)
    #     print(type(value))
    #     index += 1

    for j in range(1, max_column):

        # получение значения с выбранной ячейки, где row - строка, column - колонка
        value = worksheet.cell(row=i, column=j).value
        print(f"index: {i} {j}")
        # получение значения с выбранной ячейки, где в квадратных скобках координаты ячейки
        # value = worksheet[f"A{i}"].value

        "Almaty Taraz"

        # if value is not None:
        #     pass
        # if len(str(value)) >= 2:

        #         первый                второй
        if (value == "Almaty") or (value == "Taraz"):
            print(value)
            print(type(value))
            index += 1

print(index)
# for i in range(1, max_row):
#     for j in "ABCD":
#
#
#
# for name in var_list:
#     row = num
#     # 1 -> A, 3 -> C, 26 -> Z
#     col = get_column_letter(var_list.index(name) + 1)
#     # записываем значение в выбранную ячейку
#     worksheet[f'{col}{row}'] = str(name)


light = True
electro = False

#    первый    второй
if (light) and not (electro):
    print("Правда")
else:
    print("Ложь")

# or - если хоть один правда = правда
# and - если оба правда = правда