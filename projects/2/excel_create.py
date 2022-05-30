# устанавливаем библиотеку для работы с эксель
# pip install openpyxl
# импорт
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# создание "объекта" из библиотеки openpyxl
workbook = Workbook()

# берём активную страницу из рабочей книги
worksheet = workbook.active

var_list = ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]

# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]
# ["Almaty", "Nur-Sultan", "Taraz", "Ekibastuz"]

var_range = range(1, 1000, 1)
# print(var_range)
# for n in var_range:
#     print(n)

# функция range возвращает массив чисел
# x = range(6+1)
# for n in x:
#     # print(n)
#     pass

# x = range(3, 200, 2)
# for n in x:
#     print(n)

# "кривой способ решения задачи"
# for j in "ABCD":
#     row = "1"
#     col = j
#     # записываем значение в выбранную ячейку
#     worksheet[f'{col}{row}'] = str(var_list["ABCD".index(j)])

# пока вложенный цикл не отработает полностью, вторая итерация верхнего цикла не запускается
for num in var_range:
    for name in var_list:
        row = num
        # 1 -> A, 3 -> C, 26 -> Z
        col = get_column_letter(var_list.index(name) + 1)
        # записываем значение в выбранную ячейку
        worksheet[f'{col}{row}'] = str(name)

# записываем значение в выбранную (А1) ячейку
# worksheet['A1'] = 42

# сохраняем рабочую книгу в excel-файл(xlsx/xls)
workbook.save("sample_example.xlsx")


