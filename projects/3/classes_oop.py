import math
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

filename = 'users.xlsx'
filedir = 'home_work'

workbook = openpyxl.load_workbook(filedir + "/" + filename)
worksheet = workbook.active

max_row = worksheet.max_row
max_column = worksheet.max_column

workers = []


class Worker:  # определение (создание базовых параметров) класса
    def __init__(self, first_name: str, second_name: str, patronymic: str, position,
                 category="Рабочий"):  # магический метод для инициализации (создания) класса
        self.value_first_name = first_name  # свойство (переменная) внутри класса - изменяемая
        self.value_second_name = second_name  # свойство (переменная) внутри класса - изменяемая
        self.value_patronymic = patronymic  # свойство (переменная) внутри класса - изменяемая
        self.value_position = position  # свойство (переменная) внутри класса - изменяемая
        self.value_category = category  # свойство (переменная) внутри класса - изменяемая

        self.radius = 0

    def print_full_name(self):
        print(f'{self.value_second_name} {self.value_first_name}')

    def get_full_name(self):
        return f'{self.value_second_name} {self.value_first_name}'

    def calcuate(self, side=14):
        self.radius = side ** 2


for row in range(1, max_row + 1):
    worker = []

    for column in range(1, 5 + 1):
        obj = worksheet.cell(row=row, column=column)  # метод (функция в классе) который получает объект по координатам
        # print(obj)
        # print(type(obj))

        value = obj.value  # свойство (переменная в классе) которое получает значение ячейки

        if not value:
            value = ''
        worker.append(value)

    # print(worker)
    # print('читаю нового работника!')
    worker_obj = Worker(
        first_name=worker[1],
        second_name=worker[0],
        patronymic=worker[2],
        position=worker[3],
        category=worker[4],
    )

    # worker_obj.print_full_name()
    full_name = worker_obj.get_full_name()
    # print(full_name)

    workers.append(worker_obj)
    # workers.append(worker)
# print(workers)

print(workers[1].value_patronymic)

print(f'{workers[1].value_second_name} {workers[1].value_first_name}')  # Исагулов  Куаныш
workers[1].print_full_name()  # Исагулов  Куаныш
print(workers[1].get_full_name())  # Исагулов  Куаныш

external_value_variable = 1

workbook = Workbook()
print(workbook)
print(type(workbook))


class CustomWorkerClass1:  # определение (создание базовых параметров) класса
    def __init__(self):  # магический метод для инициализации (создания) класса
        self.VALUE_CONSTANT = 321  # свойство (переменная) внутри класса - константа (не изменяется)


print(CustomWorkerClass1)
print(type(CustomWorkerClass1))

worker = CustomWorkerClass1()  # создание экземпляра класса CustomWorkerClass
print(worker)
print(type(worker))
print(worker.VALUE_CONSTANT)  # получение свойства данного экземпляра данного класса


class CustomWorkerClass2:  # определение (создание базовых параметров) класса
    def __init__(self, external_value_variable2=15):  # магический метод для инициализации (создания) класса
        self.VALUE_CONSTANT = 321  # свойство (переменная) внутри класса - константа (не изменяется)
        self.value_variable = external_value_variable2  # свойство (переменная) внутри класса - изменяемая


worker = CustomWorkerClass2(external_value_variable2=156)  # создание экземпляра класса CustomWorkerClass
print(worker)
print(type(worker))
print(worker.VALUE_CONSTANT)  # получение свойства данного экземпляра данного класса
print(worker.value_variable)  # получение свойства данного экземпляра данного класса

print("\n\n\n**********\n\n\n")


# класс для расчёта периметров и площадей
class MyClass:
    def __init__(self, name: str, side1: float, side2: float, side3: float, side4=0.0):
        self.name = name
        self.side1 = side1
        self.side2 = side2
        self.side3 = side3
        self.side4 = side4

        self.radius = 0

    def print_name(self):
        print(self.name)

    def get_perimeter(self):
        perimeter_value = self.side1 + self.side2 + self.side3 + self.side4
        return perimeter_value

    def get_square_with_multiply(self, multiply=1.0):
        if self.side4 > 0:
            # фигуры с 4 сторонами
            return (self.side1 * self.side2) * multiply
        else:
            side1 = self.side1
            print(f'side1: {side1}')
            side2 = self.side2
            print(f'side2: {side2}')
            side3 = self.side3
            print(f'side3: {side3}')
            p = (side1 + side2 + side3) / 2
            print(p)
            value1 = p * (p - side1) * (p - side2) * (p - side3)
            print(value1)
            s = math.sqrt(value1)
            print(s)
            return s * multiply

    def calcuate(self, side=14):
        self.radius = side ** 2


figure1 = MyClass(name="квадрат", side4=10, side2=10, side3=10, side1=10)  # квадрат
figure1.print_name()

perimeter_perimeter_figure1 = figure1.get_perimeter()
print(perimeter_perimeter_figure1)

square_perimeter_figure1 = figure1.get_square_with_multiply(multiply=2.5)
print(square_perimeter_figure1)

figure2 = MyClass(name="прямоугольник", side4=20, side2=20, side3=7, side1=7)  # прямоугольник
figure2.print_name()

perimeter_perimeter_figure2 = figure2.get_perimeter()
print(perimeter_perimeter_figure2)

square_perimeter_figure2 = figure2.get_square_with_multiply(0.75)
print(square_perimeter_figure2)

figure3 = MyClass(name="треугольник", side2=20, side3=15, side1=7)  # треугольник
figure3.print_name()

perimeter_perimeter_figure3 = figure3.get_perimeter()
print(perimeter_perimeter_figure3)

square_perimeter_figure3 = figure3.get_square_with_multiply(0.75)
print(square_perimeter_figure3)

# new_data = float(input("Введите любое число: "))  # возвращает данные, которые ввёл пользователь в виде строки
# print(new_data)
# print(type(new_data))

# side1 = float(input("Введите первую сторону: "))

# user_string = input("Введите через запятую стороны объекта (12, 35, 65...): ")  # '12, 56,89, 40'
# print(user_string)
# print(type(user_string))
# # user_string = '12, 56,89, 40'
#
# # обработка ошибок
#
# sides = []
# for x in user_string.split(sep=','):
#     value = float(str(x).strip())
#     sides.append(value)
# print(sides)
#
# # sides = [x for x in]
# if len(sides) == 3:
#     sides.append(0.0)
# elif len(sides) < 3:
#     print(f"Вы ввели только {len(sides)} сторону(-ы)!")
# print(sides)
#
# try:
#     figure4 = MyClass(name="Новый объект", side1=sides[0], side2=sides[1], side3=sides[2], side4=sides[3])  # Новый объект
# except:
#     figure4 = MyClass(name="Новый объект", side1=sides[0], side2=sides[1], side3=sides[2])  # Новый объект
#
#
# figure4.print_name()
#
# perimeter_perimeter_figure4 = figure4.get_perimeter()
# print(perimeter_perimeter_figure4)
#
# square_perimeter_figure4 = figure4.get_square_with_multiply(0.5)
# print(square_perimeter_figure4)


print("\n\n\n**********\n\n\n")


class MyCalculator:
    def __init__(self, val1: float, val2: float):
        try:
            self.val1 = float(val1)
        except Exception as error:
            self.val1 = input('Введите первое значение ещё раз: ')
        self.val2 = val2

    def summ2(self):
        try:
            return float(self.val1) + float(self.val2)
        except Exception as error:
            print(error)
            return 0.0

    def multiply(self):
        return float(self.val1) * float(self.val2)

    @staticmethod
    def multiply_static(val1, val2):
        return float(val1) * float(val2)

    @staticmethod  # декоратор, который делает метод в классе статическим(без параметра селф и инициализации)
    def summ(val1: float, val2: float):  # статический метод
        return val1 + val2


# инициализация калькулятора
# summ1 = MyCalculator('sdfsd1sdg2', "16")
# print(summ1.summ2())
# print(summ1.multiply())
# print(MyCalculator.summ(15, 17))
# print(MyCalculator.multiply_static(15, 17))


# def calc_3(number1, number2, operation="-"):
#     # print(number1, number2, operation)
#     if operation == "+":
#         return number1 + number2
#     if operation == "-":
#         return number1 - number2
#     if operation == "*":
#         return number1 * number2
#     if operation == "/":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 / number2
#     if operation == "**":
#         return number1 ** number2
#     if operation == "//":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 // number2
#     if operation == "sqrt":
#         return math.sqrt(number1)
#     if operation == "%":
#         if number2 == 0:
#             print("Второе число не может быть 0")
#         else:
#             return number1 % number2

class MyTree:
    def __init__(self, name):
        self.age = 1
        self.name = name

    def drow(self, value_: float):  # рост дерева(старение)
        self.age += value_

    @staticmethod
    def drow_static(value_: float, name: str):
        age = 1
        age += value_
        # tree = {"name": name, "age": age}
        # return tree
        return dict(name=name, age=age)

tree1 = MyTree("дерево 1")  # 1
print('возраст ', tree1.name, ' = ', tree1.age)  # 1
tree1.drow(2)  # 3
print('возраст ', tree1.name, ' = ', tree1.age)  # 3
tree1.drow(2)  # 5
print('возраст ', tree1.name, ' = ', tree1.age)  # 5

tree2 = MyTree("дерево 2")  # 1
print('возраст ', tree2.name, ' = ', tree2.age)  # 1
tree2.drow(1)  # 2
print('возраст ', tree2.name, ' = ', tree2.age)  # 2

tree_dict = MyTree.drow_static(2, "дерево 3")
print(f"имя дерева: '{tree_dict['name']}', возраст дерева: '{tree_dict['age']}'")
# имя дерева: 'дерево 3', возраст дерева: '3'
