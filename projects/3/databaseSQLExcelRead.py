import psycopg2
import openpyxl

# Read all data from .xlsx
book = openpyxl.load_workbook("excel_data/tab.xlsx")
sheet = book.active
global_rows = []
for row in range(2, sheet.max_row + 1):
    local_row = []
    for column in range(1, sheet.max_column + 1):
        local_row.append(sheet.cell(row=row, column=column).value)
    global_rows.append(local_row)


# Example
# conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
# cur = conn.cursor()
# query_string = f"""
# INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
# VALUES ('tovar', 'gruppa', 'postavshick', '2022-06-06', 'region', '500', '600', 'true')
# """
# cur.execute(query_string)
# conn.commit()


# Insert all data to SQL
class Record:
    def __init__(self, row: list):
        self.tovar = str(row[0])
        self.gruppa = str(row[1])
        self.postavshick = str(row[2])
        date_post = str(row[3]).split(" ")[0]  # '2010-12-4'
        year = date_post.split('-')[0]
        month = int(date_post.split('-')[1])  # '09' -> 9
        if int(month) < 10:
            month = f"0{month}"
        day = int(date_post.split('-')[2])
        if int(day) < 10:
            day = f"0{day}"
        self.date_post = str(f"{year}-{month}-{day}")
        self.region = str(row[4])
        try:
            self.prodashi = int(row[5])
        except Exception as error:
            print(error)
            self.prodashi = 0
        try:
            self.sbit = int(row[6])
        except Exception as error:
            print(error)
            self.sbit = 0
        pribil = str(row[7])
        if pribil.lower() == "да":
            self.pribil = True
        else:
            self.pribil = False


conn = psycopg2.connect("dbname=example user=postgres password=31284bogdan")
cur = conn.cursor()
for row in global_rows:
    obj = Record(row=row)
    try:
        query_string = f"""
        INSERT INTO public.products (tovar, gruppa, postavshick, date_post, region, prodashi, sbit, pribil)
        VALUES ('{obj.tovar}', '{obj.gruppa}', '{obj.postavshick}', '{obj.date_post}', '{obj.region}', '{obj.prodashi}', 
        '{obj.sbit}', '{obj.pribil}')
        """
        cur.execute(query_string)
        conn.commit()
    except Exception as error:
        print(error)
        print(row)
conn.close()
