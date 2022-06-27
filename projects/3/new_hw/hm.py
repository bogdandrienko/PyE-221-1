import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

names = ['Лист1.xlsx', 'Лист2.xlsx', 'Лист3.xlsx']

# прочитать всю папку

global_arr = []
for filename in names:
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    max_row = worksheet.max_row + 1
    max_column = worksheet.max_column + 1
    local_arr = []
    for j in range(1, max_column):
        for i in range(1, max_row):
            value = worksheet.cell(row=i, column=j).value
            if value:
                local_arr.append(value)
    # print(local_arr)
    global_arr.append(local_arr)
print(global_arr)


workbook = Workbook()
worksheet = workbook.active

a = global_arr[0]
print(a)
a_len = len(a)
print(a_len)
a_range = range(1, a_len+1)
print(a_range)
for i in a_range:
    print(i)

b = global_arr[1]
print(b)
b_len = len(b)
print(b_len)
b_range = range(1, b_len+1)
print(b_range)
for i in b_range:
    print(i)

c = global_arr[2]
print(c)
c_len = len(c)
print(c_len)
c_range = range(1, c_len+1)
print(c_range)
for i in c_range:
    print(i)



# print(var_col)
# print(type(var_col))  # [0, 1, 2, 3, 4 ... ]
# for col in var_col:
#     print(global_arr[col])
    # row = len(global_arr[col-1])
    # print(row)
    # worksheet[f'{get_column_letter(col)}1'] = str('111111111111')
#     for name in var_list:
#         row = num
#         # 1 -> A, 3 -> C, 26 -> Z
#         col = get_column_letter(var_list.index(name) + 1)
#         worksheet[f'{col}{row}'] = str(name)
# workbook.save("sample_example.xlsx")


var_col = range(1, len(global_arr)+1, 1)
for i in var_col:
    print(f"i: {i}")
    letter = get_column_letter(i)
    print(f"letter: {letter}")
    for j in a_range:
        worksheet[f'A{j}'] = str(global_arr[0][j-1])
    # for j in b_range:
    #     print(f"j: {j}")
    # for j in c_range:
    #     print(f"j: {j}")


var_col = range(1, len(global_arr)+1, 1)
for i in var_col:
    letter = get_column_letter(i)
    var_row = range(1, len(global_arr[i-1])+1, 1)
    for j in var_row:
        worksheet[f'{letter}{j}'] = str(global_arr[i-1][j-1])

workbook.save("sample_example.xlsx")


workbook = Workbook()
worksheet = workbook.active

ws1 = workbook.create_sheet('first', 1)
for row in range(1, len(global_arr[0])+1, 1):
    ws1[f'A{row}'] = str(global_arr[0][row-1])

ws2 = workbook.create_sheet('second', 2)
for row in range(1, len(global_arr[1])+1, 1):
    ws2[f'B{row}'] = str(global_arr[1][row-1])

ws3 = workbook.create_sheet('third', 3)
for row in range(1, len(global_arr[2])+1, 1):
    ws3[f'C{row}'] = str(global_arr[2][row-1])

workbook.save("sample_example1.xlsx")
